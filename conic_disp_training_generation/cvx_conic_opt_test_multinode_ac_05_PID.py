############# Conic Programming Solver for Network Dispatch Optimization ######
###############################################################################
# The intention of this code is to setup a conic problem for optimizing the 
# unit commitment and dispatch of a network as exemplified by the Washington
# State University campus microgrid. 
#
# the script follows the sequence below:
# 1) system parameters are loaded
# 2) object classes are defined
# 3) constraint functions are defined
# 4) component variables are instantiated
# 5) objectiv function is created
# 6) problem constraint functions are collected
# 7) all variables and constraints are loaded to a problem instance
# 8) the problem definition is sent to the solver
# 9) the problem solution is sorted into desired dispatch format
#
# a microgrid parameter structure must be loaded which contains efficiency
# curves, limits, and inputs/outputs of all components, descriptions of all
# nodes, impedances of all lines, demand profiles of all nodes, and cost
# information about all resources
#
# This work is based off of the code which Kyle Monson at the Pacific
# Northwest National Lab wrote for the creation of a linear programming problem
# without a nodal network or power flow constraints.
#
# This code was written by Nadia Panossian at Washington State University and 
# was last updated on 10/17/2018 by Nadia Panossian
# the author can be reached at nadia.panossian@wsu.edu

import itertools

import os
import numpy as np
import pandas as pd
import openpyxl
import csv
from coreapi import codecs
from copy import copy
#import xlsxwriter
import cvxpy
import pickle
from class_definition.plant_struct import Plant, Network, Optimoptions
from class_definition.test_data import TestData
from class_definition.component import (ElectricChiller, AbsorptionChiller, CombinedHeatPower, ElectricGenerator, Heater)
from class_definition.component import (ElectricStorage, ThermalStorage, Utility, Renewable)
from function.setup.piecewise_fit import piecewise_quadratic, piecewise_linear
import datetime
from function.setup.update_qpform_all import (load_piecewise, fit_coproduction, remove_segments, fit_fcn)
from pickle_wsu_campus_demand import load_demand
#import os

import time




########## READ IN SYSTEM PARAMETERS
tic = time.time()
start_date = datetime.datetime(2010, 8, 1, 0, 0, 0)



#read in forecast, gen, network
with open(os.getcwd() + '\\library\\wsu_campus.pickle', 'rb') as file_object:
    plant = pickle.load(file_object)

gen = plant.generator
network = plant.network
optimoptions = plant.optimoptions


test_data = load_demand()
# with open(os.getcwd() + '\\library\\data\\wsu_campus_demand_2009_2012', 'rb') as file_object:
#     test_data = pickle.load(file_object, encoding='latin1')

## ad user inputs 
v_base = 4.135 # nominal voltage on lines kV
p_base = 5000 # power base value kVA
h_base = 20000 # heating capacity of boiler
c_base = 7.279884675000000e+03 # cooling capacity of largest chiller
current_base = p_base/v_base 
r_base = v_base**2/p_base # kohms nominal value to normalize admitance (kv^2/kVA = kohms)
j_base = r_base
v_nominal = 1.0#4135/4135
current_limit_value = 290/current_base # amps limit on 250 kcmil wire between 170 and 290
a_hru = 0.8 #system heat loss factor = 1 - a_hru, 
voltage_deviation = 0.1 # max normalized deviation of voltage
T = 24 #horizon
timesteps = 24*365 #number of receding horizon repetitions
allow_dumping= False#True
allow_thermal_slack = False
bigM = 10#1e2 #cost of not meeting demand exactly
grid_limit = 100

#functions to process information from generator list and network description
## this library sorts components by type and by type by node
# sort components into their own lists and count types
turbine_para = []
diesel_para = []
boiler_para = []
chiller_para = []
#abs_para = []
e_storage_para = []
h_storage_para = []
c_storage_para = []
grid_para = []
renew_para = []
fuel_para = []
e_storage0 = []
c_storage0 = []
h_storage0 = []
KK =5 # number of piecwise sections per component efficiency curves
turbine_pieces = []
diesel_pieces = []
boiler_pieces = []
chiller_pieces = []
#abs_pieces = []
#read in initial conditions
turbine_init = []
dieselgen_init = []
chiller_init = []
boiler_init = []
#abs_init = np.zeros((n_abs,1))
var_name_list = []

for i in range(len(gen)):
    if isinstance(gen[i], ElectricChiller):
        gen[i].size = gen[i].size/c_base
        gen[i].ramp_rate = gen[i].ramp_rate/c_base
        #create efficiency piecewise quadratic fit curve
        fit_terms,x_min, x_max = piecewise_quadratic(gen[i].output.capacity, gen[i].output.cooling*p_base/c_base, resolution=1, max_cap=gen[i].size, regression_order=3)
        setattr(gen[i], 'fundata', {"fp": fit_terms[1], "hp": fit_terms[2], "cp": fit_terms[0], "fq": 0.2*fit_terms[1], "hq": 0.2*fit_terms[2], "cq": 0.5*fit_terms[0]})
        setattr(gen[i], 'ub', x_max)
        setattr(gen[i], 'lb', x_min)
        chiller_para.append(gen[i])
        chiller_pieces.append(len(x_max))
        chiller_init.append(gen[i].size/2)
    # elif isinstance(gen[i], AbsorptionChiller):
    #     #create efficiency piecewise quadratic fit curve
    #     abs_para.append(gen[i])
    elif isinstance(gen[i], Heater):
        gen[i].size = gen[i].size/h_base
        gen[i].ramp_rate = gen[i].ramp_rate/h_base
        #create efficiency piecewise linear fit curve
        fit_terms, x_min, x_max = piecewise_quadratic(gen[i].output.capacity, gen[i].output.heat, resolution=1, max_cap=gen[i].size)
        setattr(gen[i], 'fundata', {"h": fit_terms[2], "f": fit_terms[1], "c": fit_terms[0]})
        setattr(gen[i], 'ub', x_max)
        setattr(gen[i], 'lb', x_min)
        boiler_para.append(gen[i])
        boiler_pieces.append(len(x_max))
        boiler_init.append(gen[i].size/2)
    elif isinstance(gen[i], CombinedHeatPower):
        gen[i].size = gen[i].size/p_base
        gen[i].ramp_rate = gen[i].ramp_rate/p_base
        #create efficiency piecewise quadratic fit curve for electrical output
        fit_terms, x_min, x_max = piecewise_quadratic(gen[i].output.capacity, gen[i].output.electricity, resolution=1, max_cap=gen[i].size)
        setattr(gen[i], 'fundata', {"fp": fit_terms[1], "hp": fit_terms[2], "cp": fit_terms[0], "fq": 0.2*fit_terms[1], "hq": 0.2*fit_terms[2], "cq": 0.5*fit_terms[0]})
        setattr(gen[i], 'ub', x_max)
        setattr(gen[i], 'lb', x_min)
        fit_terms,_,_ = piecewise_linear(gen[i].output.capacity, gen[i].output.heat*p_base/h_base, resolution=1, max_cap=gen[i].size)
        gen[i].fundata["f_heat"] = 0.5*p_base/h_base#fit_terms[0]
        gen[i].fundata["c_heat"] = 0#fit_terms[1]
        if gen[i].source == 'diesel':
            diesel_para.append(gen[i])
            diesel_pieces.append(len(x_max))
            dieselgen_init.append(gen[i].size/2)
        else:
            turbine_para.append(gen[i])
            turbine_pieces.append(len(x_max))
            turbine_init.append(gen[i].size/2)
    elif isinstance(gen[i], ElectricGenerator):
        gen[i].size = gen[i].size/p_base
        gen[i].ramp_rate = gen[i].ramp_rate/p_base
        #create efficiency piecewise quadratic fit curve
        fit_terms, x_min, x_max = piecewise_quadratic(gen[i].output.capacity, gen[i].output.electricity, error_thresh=0.1, resolution=1, max_cap=gen[i].size)
        setattr(gen[i], 'ub', x_max)
        setattr(gen[i], 'lb', x_min)
        setattr(gen[i], 'fundata', {"fp": fit_terms[1], "hp": fit_terms[2], "cp": fit_terms[0], "fq": 0.2*fit_terms[1], "hq": 0.2*fit_terms[2], "cq": 0.5*fit_terms[0]})
        if gen[i].source == 'diesel':
            diesel_para.append(gen[i])
            diesel_pieces.append(len(x_max))
            dieselgen_init.append(gen[i].size/2)
        else:
            turbine_para.append(gen[i])
            turbine_pieces.append(len(x_max))
            turbine_init.append(gen[i].size/2)
    elif isinstance(gen[i], ElectricStorage):
        gen[i].size = gen[i].size/p_base
        gen[i].ramp_rate = gen[i].ramp_rate/p_base
        e_storage_para.append(gen[i])
        e_storage0.append(gen[i].size/2)
    elif isinstance(gen[i], ThermalStorage) and gen[i].source == 'heat':
        gen[i].size = gen[i].size/h_base
        gen[i].ramp_rate = gen[i].ramp_rate/h_base
        h_storage_para.append(gen[i])
        h_storage0.append(gen[i].size/2)
    elif isinstance(gen[i], ThermalStorage) and gen[i].source == 'cooling':
        gen[i].size = gen[i].size/c_base
        gen[i].ramp_rate = gen[i].ramp_rate/c_base
        c_storage_para.append(gen[i])
        c_storage0.append(gen[i].size/2)
    elif isinstance(gen[i], Utility) and gen[i].source == 'electricity':
        gen[i].size = gen[i].size/p_base
        grid_para.append(gen[i])
    elif isinstance(gen[i], Utility):
        fuel_para.append(gen[i])
    elif isinstance(gen[i], Renewable):
        gen[i].size = gen[i].size/p_base
        renew_para.append(gen[i])

# convert_quadratic is a function which takes a quadratic from the
# form (hx^2 + fx + c) to the format (bx + c)^2 + ex + d
def convert_quadratic(gen_para):
    j = 0
    for gt in gen_para:
        if len(gt.fundata) == 3:
            h = gt.fundata["h"]
            f = gt.fundata["f"]
            c = gt.fundata["c"]
            h[h<0] = 0
            b = np.sqrt(h)
            ci = np.divide(f,2*b)
            e = np.zeros((len(f),))
            e[np.isinf(ci)] = f[np.isinf(ci)]
            ci[np.isinf(ci)] = 0
            d = c - sum(np.power(ci,2))
            gen_para[j].fundata["b"] = b
            gen_para[j].fundata["ci"] = ci
            gen_para[j].fundata["d"] = d
            gen_para[j].fundata["e"] = e
        else:
            hp = gt.fundata["hp"]
            hq = gt.fundata["hq"]
            fp = gt.fundata["fp"]
            fq = gt.fundata["fq"]
            cp = gt.fundata["cp"]
            cq = gt.fundata["cq"]
            #filter out rounding errors to make sure curvature is positive
            hp[hp<0] = 0
            hq[hq<0] = 0
            #convert to (bx + c)^2 + ex + d 
            bp = np.sqrt(hp)
            bq = np.sqrt(hq)
            cip = np.divide(fp,2*bp)
            ciq = np.divide(fp,2*bq)
            ep = np.zeros((len(fp),))
            eq = np.zeros((len(fq),))
            ep[np.isinf(cip)] = fp[np.isinf(cip)]
            eq[np.isinf(ciq)] = fq[np.isinf(ciq)]
            cip[np.isinf(cip)] = 0
            ciq[np.isinf(ciq)] = 0
            dp = cp - sum(np.power(cip,2))
            dq = cq - sum(np.power(ciq,2))
            #save information to parameters structure
            gen_para[j].fundata["bp"] = bp
            gen_para[j].fundata["cip"] = cip
            gen_para[j].fundata["dp"] = dp
            gen_para[j].fundata["bq"] = bq
            gen_para[j].fundata["ciq"] = ciq
            gen_para[j].fundata["dq"] = dq
            gen_para[j].fundata["ep"] = ep
            gen_para[j].fundata["eq"] = eq
        j +=1
    return gen_para
    # b = sqrt(h)
    # ci = f/2/b
    # d = c - ci^2
    # return b, ci, d

#convert piecewise quadratic fit curves from format (hx^2 + fx + c) to format (bx + c)^2 + d or (hx)^2 + fx + c
turbine_para = convert_quadratic(turbine_para)
diesel_para = convert_quadratic(diesel_para)
chiller_para = convert_quadratic(chiller_para)
boiler_para = convert_quadratic(boiler_para)
    


#find_nodes creates a list of lists of indexes of equipment by node
def find_nodes(comp_list):
    comp_by_node = []
    for node in network:
        this_node = []
        i = 0
        for comp in comp_list:
            check = any([True for equip in node.equipment if equip.name==comp.name])
            if check:
                this_node.append(i)
            i = i+1
        comp_by_node.append(this_node)
    return comp_by_node

#create a list of lines by node
e_lines_by_node = [] #list of connections by list of nodes
c_lines_by_node = []
h_lines_by_node = []
i_e = 0 #index of line connection
i_c = 0
i_h = 0
i = 0
e_nodes = [] #list of electrical nodes
e_connected_nodes = []
c_nodes = []
h_nodes = []
for node in network:
    n_e = len(node.electrical.connections)
    n_c = len(node.district_cooling.connections)
    n_h = len(node.district_heat.connections)
    e_lines_by_node.append([i_e+i for i in range(n_e)])
    c_lines_by_node.append([i_c+i for i in range(n_c)])
    h_lines_by_node.append([i_h+i for i in range(n_h)])
    i_e += n_e
    i_c += n_c
    i_h += n_h
    if n_e>0:
        e_connected_nodes.append(i)
    if n_e>0 or node.electrical.load !=None:
        e_nodes.append(i)
    if n_c>0 or node.district_cooling.load != None:
        c_nodes.append(i)
    if n_h>0 or node.district_heat.load !=None:
        h_nodes.append(i)
    i +=1
n_e_nodes = len(e_nodes) #number of electrical nodes
n_e_connected_nodes = len(e_connected_nodes) #number of electrical nodes that are connected to other nodes
n_c_nodes = len(c_nodes)
n_h_nodes = len(h_nodes)
n_e_lines = i_e
n_c_lines = i_c
n_h_lines = i_h

#G = np.ones((len(network), len(network)))*-4
#B = np.ones((len(network), len(network)))*-6
# there are two types of distribution lines on campus: 250KCMil and 350KCMIL
Z_250_real = 0.240440 #/1000 #real portion of impedance in kOhms per mile
Z_250_im = 0.167776 #/1000 
Z_350_real = 0.248339 #/1000
Z_350_im = 0.173504 #/1000
#normalize them
Z_250_real = Z_250_real/r_base
Z_250_im = Z_250_im/j_base
Z_350_real = Z_350_real/r_base
Z_350_im = Z_350_im/j_base

# line lengths are described on the one line diagram from Nathan
SPU125_to_chillers = Z_250_real*600 *(1/5280) + Z_250_im*600j *(1/5280)# impedance per mile * feet * miles/feet
TUR115_and_SPU125_to_WChillers = Z_250_real*(600)*(1/5280) + Z_250_im*(600j)*(1/5280) 
#SPU125_to_SPU122_real = Z_250_real*(68+2070+820+370+2510)*(1/5280)
#SPU125_to_SPU122_im = Z_250_im*(68j+2070j+820j+370j+2510j)*(1/5280)
SPU122_to_WCampusBuild = Z_250_real*(2510+ (2070+820+370)/2)*(1/5280) + Z_250_im*(2510j+ (2070j+820j+370j)/2)*(1/5280)
SPU122_to_GT2 = Z_350_real*654*(1/5280) + Z_350_im*654j*(1/5280)
SPU124_to_GT3 = Z_250_im*830j*(1/5280) + Z_250_real*830*(1/5280)
#SPU124_to_CCampusBuild_real = Z_250_real*(924+667/2)*(1/5280)
#SPU124_to_CCampusBuild_im = Z_250_im*(924j+667j/2)*(1/5280)
SPU124_to_JohnsonHall = Z_250_real*(245+667+924)*(1/5280) + Z_250_im*(245j+667j+924j)*(1/5280)
# these values are made up
TUR117_to_PV = Z_250_im*(600j)*(1/5280) + Z_250_real*(600)*(1/5280)
TUR111_to_SCUE = Z_250_im*(1320j)*(1/5280) + Z_250_real*(1320)*(1/5280)
TVW131_to_EChillers = Z_250_real*1320*(1/5280)+Z_250_im*1320j*(1/5280)
TUR115_to_SPU125 = Z_250_real*100*(1/5280) + Z_250_im*100j*(1/5280)

G = np.zeros((len(network), len(network)))
B = np.zeros((len(network), len(network)))
Z = np.zeros((len(network), len(network)), dtype=complex)
Z[0,0] = TUR115_and_SPU125_to_WChillers
Z[0,1] = TUR115_to_SPU125
Z[1,0] = Z[0,1]
Z[1,1] = TUR115_and_SPU125_to_WChillers
Z[2,2] = 1/(1/SPU122_to_GT2 + 1/SPU122_to_WCampusBuild)
Z[2,3] = SPU122_to_GT2
Z[3,2] = Z[2,3]
Z[3,3] = 1/(1/SPU124_to_GT3 + 1/SPU124_to_JohnsonHall)
Z[4,4] = TUR111_to_SCUE
Z[5,5] = TVW131_to_EChillers
Z[6,6] = TUR117_to_PV
# Z[0,0] = (1/(TUR115_and_SPU125_to_WChillers_im + TUR115_and_SPU125_to_WChillers_real) + (1/TUR115_to_SPU125))
# Z[0,1] = -(1/TUR115_to_SPU125)
# Z[1,0] = Z[0,1]
# Z[1,1] = (1/(TUR115_and_SPU125_to_WChillers_im + TUR115_and_SPU125_to_WChillers_real) + (1/TUR115_to_SPU125))
# Z[2,2] = 1/(SPU122_to_WCampusBuild_im + SPU122_to_WCampusBuild_real) + (1/(SPU122_to_GT2_im + SPU122_to_GT2_real))
# Z[2,3] = -(1/(SPU122_to_GT2_im + SPU122_to_GT2_real))
# Z[3,2] = Z[2,3]
# Z[3,3] = (1/(SPU124_to_CCampusBuild_im + SPU124_to_CCampusBuild_real) + 1/(SPU124_to_JohnsonHall_im + SPU124_to_JohnsonHall_real)) + (1/(SPU122_to_GT2_im + SPU122_to_GT2_real))
# Z[4,4] = 1/TUR111_to_SCUE
# Z[5,5] = 1/TVW131_to_EChillers
# Z[6,6] = 1/TUR117_to_PV

Z = np.linalg.inv(Z)
G = Z.real
B = Z.imag

# Q_loss = (6.8754E-10 Q_supply + 0.0056345)*L
cool_a = 6.8754E-10
cool_b = 0.0056345/c_base
cool_loss = np.zeros((len(network), len(network)))
cool_loss[1,2] = 1
cool_loss[2,5] = 1
cool_loss[5,3] = 1
cool_loss[3,4] = 1
cool_loss[4,1] = 1
cool_loss[2,1] = 850*(cool_a)-1
cool_loss[5,2] = 1100*(cool_a)-1
cool_loss[3,5] = 400*(cool_a)-1
cool_loss[4,3] = 550*(cool_a)-1
cool_loss[1,4] = 800*(cool_a)-1

cool_loss_const = np.zeros((len(network),1))
cool_loss_const[1] = 800*cool_b
cool_loss_const[2] = 850*cool_b
cool_loss_const[5] = 1100*cool_b
cool_loss_const[3] = 400*cool_b
cool_loss_const[4] = 550*cool_b
# cool_loss[0,1] = 1
# cool_loss[1,2] = 1
# cool_loss[2,3] = 1
# cool_loss[3,4] = 1
# cool_loss[4,5] = 1
# cool_loss[5,0] = 1
# #cool_loss[5,6] = 1
# #cool_loss[6,0] = 1
# #cool_loss[0,6] = -0.95
# cool_loss[0,5] = -0.95
# cool_loss[1,0] = -0.95
# cool_loss[2,1] = -0.95
# cool_loss[3,2] = -0.95
# cool_loss[4,3] = -0.95
# cool_loss[5,4] = -0.95
#cool_loss[6,5] = -0.9

# Q_loss = (1.4077E-9 Q_supply + 0.062685)*L
heat_a = 1.4077E-9
heat_b = 0.062685/h_base
heat_loss = np.zeros((len(network), len(network)))
heat_loss[0,4] = 1
heat_loss[4,2] = 1
heat_loss[2,3] = 1
heat_loss[3,5] = 1
heat_loss[5,0] = 1
heat_loss[4,0] = 850*heat_a-1
heat_loss[2,4] = 900*heat_a-1
heat_loss[3,2] = 850*heat_a-1
heat_loss[5,3] = 250*heat_a-1
heat_loss[0,5] = 1500*heat_a-1

heat_loss_const = np.zeros((len(network),1))
heat_loss_const[0] = 1500*heat_b
heat_loss_const[2] = 900*heat_b
heat_loss_const[3] = 850*heat_b
heat_loss_const[4] = 850*heat_b
heat_loss_const[5] = 250*heat_b
# heat_loss[0,1] = 1
# heat_loss[1,2] = 1
# heat_loss[2,3] = 1
# heat_loss[3,4] = 1
# heat_loss[4,5] = 1
# heat_loss[5,0] = 1
# #heat_loss[5,6] = 1
# #heat_loss[6,0] = 1
# #heat_loss[0,6] = -0.95
# heat_loss[0,5] = -0.95
# heat_loss[1,0] = -0.95
# heat_loss[2,1] = -0.95
# heat_loss[3,2] = -0.95
# heat_loss[4,3] = -0.95
# heat_loss[5,4] = -0.95
#heat_loss[6,5] = -0.95


n_turbines = len(turbine_para)
n_dieselgen = len(diesel_para)
n_boilers = len(boiler_para)
n_chillers = len(chiller_para)
#n_abs = len(abs_para) # zero absorption chillers
n_egrid = 1
n_e_storage = len(e_storage_para)
n_c_storage = len(c_storage_para)
n_h_storage = len(h_storage_para)
n_components = len(gen)
n_nodes = len(network)
states = []
constraints = []
date_range = [start_date+datetime.timedelta(hours=i) for i in range(T)]
## create list of components of each type at each node
grid_by_node = find_nodes(grid_para)
turbine_by_node = find_nodes(turbine_para)
diesel_by_node = find_nodes(diesel_para)
boiler_by_node = find_nodes(boiler_para)
chiller_by_node = find_nodes(chiller_para)
#abs_by_node = find_nodes(abs_para)
e_storage_by_node = find_nodes(e_storage_para)
h_storage_by_node = find_nodes(h_storage_para)
c_storage_by_node = find_nodes(c_storage_para)
renew_by_node = find_nodes(renew_para)

# set the starting value for x_n
# it is used as an upper limit, so allow it to be the upper bound on voltage
x_n = np.ones((n_e_nodes,T))*(1+voltage_deviation)**2


# variable to indicate that we want all variables that match a pattern
# one item in the tuple key can be RANGE
RANGE = -1


####### POWER NETWORK OBJECT CLASS DEFINITIONS AND FUNCTIONS
class BuildAsset(object):
    def __init__(self, fundata, ramp_up=None, ramp_down=None, startcost=0, min_on=0, min_off=0, component_name=None):
        self.component_name = component_name
        self.fundata = fundata
        self.ramp_up = ramp_up
        self.ramp_down = ramp_down
        self.startcost = startcost
        self.min_on = min_on
        self.min_off = min_off

class BuildAsset_init(object):
    def __init__(self, status=0, output=0.0):
        self.status = status
        self.status1 = np.zeros(T)
        self.output = output

#  storage class
class Storage(object):
    def __init__(self, pmax, Emax, eta_ch, eta_disch, soc_max=1.0, soc_min=0.0, now_soc=0.0, component_name=None):
        self.pmax = pmax
        self.Emax = Emax
        self.eta_ch = eta_ch
        self.eta_disch = eta_disch
        self.soc_max = soc_max
        self.soc_min = soc_min
        self.now_soc = now_soc
        self.component_name = component_name
        if now_soc is None:
            raise ValueError("STATE OF CHARGE IS NONE")

def binary_var(var_name):
    return cvxpy.Variable(name=var_name, boolean=True)

#forecast function
def find_demand(date_stamp,demand_type, n=0):
    if n != None:
        f_ind = test_data.timestamp.index(date_stamp)
        if demand_type == 'e':
            demand = getattr(test_data.demand,demand_type)[n][f_ind]/p_base
        elif demand_type == 'h':
            demand = getattr(test_data.demand,demand_type)[n][f_ind]/h_base
        elif demand_type == 'c':
            demand = getattr(test_data.demand,demand_type)[n][f_ind]/c_base
    else:
        demand = 0.0
    return demand

def find_solar_forecast(date_stamp, n=0):
    f_ind = test_data.timestamp.index(date_stamp)
    irrad = test_data.weather.irrad_dire_norm[f_ind]
    if irrad == -9900: #this value is an error from the sensor
        irrad = 0
    solar_gen = 0
    if not renew_by_node[n] == []:
        solar_gen = np.abs(sum([irrad * (renew_para[i].size_m2*renew_para[i].gen_frac)/1000 for i in renew_by_node[n]]))/p_base
    return solar_gen


#electric utility pricing function
def find_utility_pricing(date_stamp):
    i = 0
    weekday = date_stamp.weekday()
    hour = date_stamp.hour
    month = date_stamp.month
    day = date_stamp.day
    sum_start_month = grid_para[i].sum_start_month
    win_start_month = grid_para[i].win_start_month
    sum_start_day = grid_para[i].sum_start_day
    win_start_day = grid_para[i].win_start_day
    #determine which rate table to use (summer or winter)
    # then apply the indices from the rate table to the find the actual rates
    if month > sum_start_month and month < win_start_month:
        i_rate = grid_para[i].sum_rate_table[weekday][hour]
        rate = grid_para[i].sum_rates[i_rate,0]
    elif month < sum_start_month or month > win_start_month:
        i_rate = grid_para[i].win_rate_table[weekday][hour]
        rate = grid_para[i].win_rates[i_rate,0]
    elif month == sum_start_month:
        if day>= sum_start_day:
            i_rate = grid_para[i].sum_rate_table[weekday][hour]
            rate = grid_para[i].sum_rates[i_rate,0]
        else:
            i_rate = grid_para[i].win_rate_table[weekday][hour]
            rate = grid_para[i].win_rates[i_rate,0]
    elif month == win_start_month:
        if day>= win_start_day:
            i_rate = grid_para[i].win_rate_table[weekday][hour]
            rate = grid_para[i].win_rates[i_rate,0]
        else:
            i_rate = grid_para[i].sum_rate_table[weekday][hour]
            rate = grid_para[i].sum_rates[i_rate,0]
    return rate

#gas utility pricing function
def find_gas_pricing(date_stamp):
    i=0
    day_stamp = datetime.datetime(year=date_stamp.year, month=date_stamp.month, day=date_stamp.day)
    price_ind = fuel_para[i].timestamp.index(day_stamp)
    gas_rate = fuel_para[i].rate[price_ind] # price in $/thousand cubic feet
    gas_rate = gas_rate*(1/293.07) # convert to $/kWh gas --> 293.07 kWh/thousand cubic feet natural gas
    return gas_rate

#diesel supply pricing function
def find_diesel_pricing(date_stamp):
    i=1
    day_stamp = datetime.datetime(year=date_stamp.year, month=date_stamp.month, day=date_stamp.day)
    price_ind = fuel_para[i].timestamp.index(day_stamp)
    return fuel_para[i].rate[price_ind]

#  all network objects create a group of variables associated with that object
class VariableGroup(object):
    def __init__(self, name, indexes=(), is_binary_var=False, lower_bound_func=None, upper_bound_func=None, T=T, pieces=[1]):
        global var_name_list
        self.variables = {}

        name_base = name
        #if it is a piecewise function, make the variable group be a group of arrays (1,KK)
        if pieces==[1]:
            pieces = [1 for i in indexes[0]]

        #create name base string
        for _ in range(len(indexes)):
            name_base += "_{}"

        #create variable for each timestep and each component with a corresponding name
        for index in itertools.product(*indexes):
            var_name = name_base.format(*index)

            if is_binary_var:
                var = binary_var(var_name)
            else:
                #assign upper and lower bounds for the variable
                if lower_bound_func is not None:
                    lower_bound = lower_bound_func(index)
                else:
                    lower_bound = None

                if upper_bound_func is not None:
                    upper_bound = upper_bound_func(index)
                else:
                    upper_bound = None

                #the lower bound should always be set if the upper bound is set
                if lower_bound is None and upper_bound is not None:
                    raise RuntimeError("Lower bound should not be unset while upper bound is set")

                #create the cp variable
                if lower_bound_func == constant_zero:
                    var = cvxpy.Variable(pieces[index[0]], name = var_name, nonneg=True)
                elif lower_bound is not None:
                    var = cvxpy.Variable(pieces[index[0]], name=var_name)
                    #constr = [var>=lower_bound]
                elif upper_bound is not None:
                    var = cvxpy.Variable(pieces[index[0]],name=var_name)
                    #constr = [var<=upper_bound, var>=lower_bound]
                else:
                    var = cvxpy.Variable(pieces[index[0]], name=var_name)
                

            self.variables[index] = var
            var_name_list.append(var_name)
            #self.constraints[index] = constr
        
    #internal function to find variables associated with your key
    def match(self, key):
        position = key.index(RANGE)
        def predicate(xs, ys):
            z=0
            for i, (x, y) in enumerate(zip(xs, ys)):
                if i != position and x==y:
                    z += 1
            return z == len(key)-1


        keys = list(self.variables.keys())
        keys = [k for k in keys if predicate(k,key)]
        keys.sort(key=lambda k: k[position]) 

        return [self.variables[k] for k in keys]

    #variable function to get the variables associated with the key
    def __getitem__(self, key):
        if type(key) != tuple:
            key = (key,)
            
        #n_ones = 0
        #for i, x in enumerate(key):
        #    if x == RANGE:
        #        n_ones+=1
        n_range = key.count(RANGE)

        if n_range == 0:
            return self.variables[key]
        elif n_range ==1:
            return self.match(key)
        else:
            raise ValueError("Can only get RANGE for one index.")


def constant(x):
    def _constant(*args, **kwargs):
        return x
    return _constant

constant_zero = constant(0)


######### CONSTRAINT FUNCTIONS
def add_constraint(name, indexes, constraint_func):
    name_base = name
    for _ in range(len(indexes)):
        name_base +="_{}"

    for index in itertools.product(*indexes):
        name = name_base.format(*index)
        c = constraint_func(index)
        constraints.append((c,name))

toc = time.time()-tic
print('load all data and functions' + str(toc))

def run_horizon(timestep, v_iters, x_n, pid_error_last):
    # INPUTS:
    # timestep: the timestamp for the timesteps in the horizon
    # v_iters: an integer denoting how many times you have tried to converge on this problem
    # x_n: last iterations attempt at voltage values
    # pid_error: iteration error used for pid convergence
    #
    # OUTPUTS:
    # v_iters: integer denoting either number of iterations past, or 0 indicating convergence
    # pid_error: error in voltage values as applied to the PID convergence
    # x_n: the iteration's values for voltage at nodes
    # 
    #balance equations
    #electric nodal balance
    def electric_p_balance(index):
        t, m = index
        n = [i for i in range(n_nodes) if network[i].name in network[m].electrical.connections]
        i_turb = turbine_by_node[m]
        i_grid = grid_by_node[m]
        i_chiller = chiller_by_node[m]
        i_es = e_storage_by_node[m]
        i_dies = diesel_by_node[m]
        i_renew = renew_by_node[m]
        i_lines = e_lines_by_node[m]
        #sum of power at node = Gmmxm + sum(Gmnymn+Bmnymn)
        #start with sum of power at node
        if not n== []:
            return cvxpy.sum([turbine_xp[j,t] for j in i_turb])\
            + cvxpy.sum([ep_elecfromgrid[j,t] - ep_electogrid[j,t] for j in i_grid])\
            - cvxpy.sum([chiller_yp[j,t] for j in i_chiller])\
            + cvxpy.sum([e_storage_dish[j,t] - e_storage_ch[j,t] for j in i_es])\
            + cvxpy.sum([dieselgen_xp[j,t] for j in i_dies])\
            - forecast.demand.ep[m,t]\
            + forecast.renew[m,t]\
            + elec_unserve[m,t]\
            == G[m,m]*x_m[m,t]\
            + cvxpy.sum(G[m,n[RANGE]]*y_mn[i_lines[RANGE],t] + B[m,n[RANGE]]*z_mn[i_lines[RANGE],t])
        else:
            return cvxpy.sum([turbine_xp[j,t] for j in i_turb])\
            + cvxpy.sum([ep_elecfromgrid[j,t] - ep_electogrid[j,t] for j in i_grid])\
            - cvxpy.sum([chiller_yp[j,t] for j in i_chiller])\
            + cvxpy.sum([e_storage_dish[j,t] - e_storage_ch[j,t] for j in i_es])\
            + cvxpy.sum([dieselgen_xp[j,t] for j in i_dies])\
            - forecast.demand.ep[m,t]\
            + forecast.renew[m,t]\
            + elec_unserve[m,t]\
            == G[m,m]*x_m[m,t]
        #normalized by 3000

    # reactive power nodal balance
    def electric_q_balance(index):
        t, m = index
        n = [i for i in range(n_nodes) if network[i].name in network[m].electrical.connections]
        i_turb = turbine_by_node[m]
        i_grid = grid_by_node[m]
        i_chiller = chiller_by_node[m]
        i_es = e_storage_by_node[m]
        i_dies = diesel_by_node[m]
        i_lines = e_lines_by_node[m]
        # sum of reactive power at node = -Bmmxm + sum(Gmnzmn - Bmnymn)
        # reative power at node is the sum of all power produced at that node minus any power consumed there
        # energy storage is assumed to store real power only, so not included here
        if not n==[]:
            return cvxpy.sum([turbine_xq[j,t] for j in i_turb])\
            + cvxpy.sum([eq_elecfromgrid[j,t] - eq_electogrid[j,t] for j in i_grid])\
            - cvxpy.sum([chiller_yq[j,t] for j in i_chiller])\
            + cvxpy.sum([dieselgen_xq[j,t] for j in i_dies])\
            - forecast.demand.eq[m,t]\
            == cvxpy.sum(G[m,n[RANGE]]*z_mn[i_lines[RANGE],t] - B[m,n[RANGE]]*y_mn[i_lines[RANGE],t]) -B[m,m]*x_m[m,t]
        else:
            return cvxpy.sum([turbine_xq[j,t] for j in i_turb])\
            + cvxpy.sum([eq_elecfromgrid[j,t] - eq_electogrid[j,t] for j in i_grid])\
            - cvxpy.sum([chiller_yq[j,t] for j in i_chiller])\
            + cvxpy.sum([dieselgen_xq[j,t] for j in i_dies])\
            - forecast.demand.eq[m,t]\
            == -B[m,m]*x_m[m,t]


    #heat nodal balance
    def heat_balance(index):
        t, m = index
        n = [i for i in range(n_nodes) if network[i].name in network[m].district_heat.connections]
        i_turb = turbine_by_node[m]
        i_boiler = boiler_by_node[m]
        i_hs = h_storage_by_node[m]
        i_lines = h_lines_by_node[m]
        #i_abs = abs_by_node[m]
        #sum of heat produced-heat used at this node = heat in/out of this node
        return cvxpy.sum([boiler_x[j,t] for j in i_boiler])\
        + cvxpy.sum([turbine_para[j].fundata["f_heat"]*turbine_xp[j,t] + turbine_para[j].fundata["c_heat"] for j in i_turb])\
        + cvxpy.sum([h_storage_disch[j,t] - h_storage_ch[j,t] for j in i_hs])\
        - forecast.demand.h[m,t]\
        - heat_loss_const[m][0]\
        - heat_dump[m,t]\
        - cvxpy.sum([heat_loss[m,n[j]]*h_mn[i_lines[j],t] for j in range(len(n))])\
        ==0
        # - cvxpy.sum([abs_y[j,t] for j in i_abs])\

    # cooling power nodal balance
    def cool_balance(index):
        t, m = index
        n = [i for i in range(n_nodes) if network[i].name in network[m].district_cooling.connections]
        i_chiller = chiller_by_node[m]
        #i_abs = abs_by_node[m]
        i_cs = c_storage_by_node[m]
        i_lines = c_lines_by_node[m]
        return cvxpy.sum([chiller_x[j,t] for j in i_chiller])\
        + cvxpy.sum([c_storage_disch[j,t] - c_storage_ch[j,t] for j in i_cs])\
        - forecast.demand.c[m,t]\
        - cool_loss_const[m][0]\
        - cvxpy.sum([cool_loss[m,n[j]]*c_mn[i_lines[j],t] for j in range(len(n))])\
        == 0
        #cvxpy.sum([abs_x[j,t] for j in i_abs]) + 
        #- cool_dump[m,t]\

    # voltage constraints
    def voltage_limit_lower(index):
        t, m = index
        return (v_nominal*(1.0-voltage_deviation))**2 <= x_m[m,t] 

    def voltage_limit_upper(index):
        t, m = index
        return x_m[m,t] <= (v_nominal*(voltage_deviation+1.0))**2

    def y_mn_equality(index):
        t, m = index
        # find lines out
        m_lines = e_lines_by_node[m]
        # find lines back in
        node_ns = [i for i in range(n_nodes) if network[i].name in network[m].electrical.connections]
        n_lines = []
        for n in node_ns:
            this_line = [i for i in range(len(e_lines_by_node[n])) if network[m].name == network[n].electrical.connections[i]]
            n_lines.append(e_lines_by_node[n][this_line[0]])
        # set those lines equal for y_mn
        return y_mn[m_lines[RANGE],t] == y_mn[n_lines[RANGE],t]

    def z_mn_inequality(index):
        t, m = index
        # find lines out
        m_lines = e_lines_by_node[m]
        # find lines back in
        node_ns = [i for i in range(n_nodes) if network[i].name in network[m].electrical.connections]
        n_lines = []
        for n in node_ns:
            this_line = [i for i in range(len(e_lines_by_node[n])) if network[m].name == network[n].electrical.connections[i]]
            n_lines.append(e_lines_by_node[n][this_line[0]])
        # set those lines equal to the negative of each other for z_mn
        return z_mn[m_lines[RANGE],t] == -z_mn[n_lines[RANGE],t]


    # def y_voltage_limit_lower(index):
    #     t, m = index
    #     return 0.8*(v_nominal*0.9)**2 <= y_mn[m,t]

    # def y_voltage_limit_upper(index):
    #     t, m = index 
    #     return y_mn[m,t] <= (v_nominal*0.1)**2

    # def z_voltage_limit_upper(index):
    #     t, m = index
    #     return z_mn[m,t] <= (v_nominal*100)**2

    # def z_voltage_limit_lower(index):
    #     t, m = index
    #     return -0.2*(v_nominal*1.1)**2 <= z_mn[m,t]

    # are you importing or exporting power
    def power_import_state(index):
        t, m = index
        return ep_gridstate[m,t]*grid_limit >= ep_elecfromgrid[m,t]
    def power_export_state(index):
        t, m = index
        return (1-ep_gridstate[m,t])*grid_limit >= ep_electogrid[m,t]

    # line current limits
    def current_limit(index):
        t, m = index
        n = [i for i in range(n_nodes) if network[i].name in network[m].electrical.connections]
        i_lines = e_lines_by_node[m]
        return (G[m,m]**2 + B[m,n[RANGE]]**2)*(x_m[m,t] + x_m[n[RANGE],t] - 2*y_mn[i_lines[RANGE],t])<= current_limit_value**2

    # equality to assure that x,y,z variable subsitution holds: xmxn = ymn^2 + zmn^2
    def electric_interrelation(index):
        t, m = index #m is the node
        n = [i for i in range(n_nodes) if network[i].name in network[m].electrical.connections]
        i_lines = e_lines_by_node[m]
        #return cvxpy.power(x_m[n[RANGE],t] + x_m[m,t],2) - cvxpy.power(x_m[n[RANGE],t], 2) - cvxpy.power(x_m[m,t], 2) >= cvxpy.power(y_mn[i_lines[RANGE],t], 2) + cvxpy.power(z_mn[i_lines[RANGE], t], 2) 
        #return x_m[n[RANGE],t]*x_m[m,t] == y_mn[i_lines[RANGE],t]**2 + z_mn[i_lines[RANGE],t]**2
        #return -1/2*cvxpy.square(x_m[m,t] + x_m[n[RANGE],t])+v_nominal**2 >= cvxpy.square(y_mn[i_lines[RANGE],t]) + cvxpy.square(z_mn[i_lines[RANGE],t])
        #return x_m[m,t]*x_m[n[0],t] >= y_mn[i_lines[0],t]**2 + z_mn[i_lines[0],t]**2
        
        return x_m[m,t]*x_n[n[RANGE],t] >=y_mn[i_lines[RANGE],t]**2 + z_mn[i_lines[RANGE],t]**2
        #use binomial approximation
        #return 2*x_m[m,t]-1 >=y_mn[i_lines[RANGE],t]**2 + z_mn[i_lines[RANGE],t]**2

    # equality to assure that cooling leaving one node is recorded negative from one node and positive for the other
    def line_heat(index):
        t, m = index #m is the node
        m_line = h_lines_by_node[m]
        n = [i for i in range(n_nodes) if network[i].name in network[m].district_heat.connections]
        n_line = []
        for i in n:
            i_n = network[i].district_heat.connections.index(network[m].name)
            if i_n == 0:
                n_line.append(h_lines_by_node[i][i_n])
            elif i_n ==1:
                n_line.insert(0,h_lines_by_node[i][i_n])
        return h_mn[m_line[RANGE],t] == h_mn[n_line[RANGE],t]

    def line_cooling(index):
        t, m = index #m is the node
        m_line = c_lines_by_node[m]#index of line from m to n
        n = [i for i in range(n_nodes) if network[i].name in network[m].district_cooling.connections]
        n_line = []
        for i in n:
            i_n = network[i].district_cooling.connections.index(network[m].name)
            if i_n == 0:
                n_line.append(c_lines_by_node[i][i_n])
            elif i_n ==1:
                n_line.insert(0,c_lines_by_node[i][i_n])
        return c_mn[m_line[RANGE],t] == c_mn[n_line[RANGE],t]

    # storage constraint functions

    def e_storage_state_constraint(index):
        i, t = index
        return e_storage_state[i,t] == e_storage_state[i,t-1] + e_storage_para[i].eta_ch * e_storage_ch[i,t] - 1/e_storage_para[i].eta_disch * e_storage_disch[i,t]

    def e_storage_init(index):
        i = index[0]
        return e_storage_state[i,0] == e_storage0[i] + e_storage_para[i].eta_ch * e_storage_ch[i,0] - 1/e_storage_para[i].eta_disch * e_storage_disch[i,0]

    def e_storage_end(index):
        i = index[0]
        return e_storage_state[i,T-1] == e_storage0[i]

    def h_storage_state_constraint(index):
        i, t = index
        return h_storage_state[i,t] == h_storage_state[i,t-1] + h_storage_para[i].eta_ch * h_storage_ch[i,t] - 1/h_storage_para[i].eta_disch * h_storage_disch[i,t]

    def h_storage_init(index):
        i = index[0]
        return h_storage_state[i,0] == h_storage0[i] + h_storage_para[i].eta_ch * h_storage_ch[i,0] - 1/h_storage_para[i].eta_disch * h_storage_disch[i,0]

    def h_storage_end(index):
        i = index[0]
        return h_storage_state[i,T-1] == h_storage0[i]

    def c_storage_init(index):
        i = index[0]
        return c_storage_state[i,0] == c_storage0[i] + c_storage_para[i].charge_eff * c_storage_ch[i,1] - 1/c_storage_para[i].disch_eff * c_storage_disch[i,1]

    def c_storage_state_constraint(index):
        i, t = index
        return c_storage_state[i,t] == c_storage_state[i,t-1] + c_storage_para[i].charge_eff * c_storage_ch[i,t] - 1/c_storage_para[i].disch_eff * c_storage_disch[i,t]

    def c_storage_end(index):
        i = index[0]
        return c_storage_state[i,T-1] == c_storage0[i]

    # turbnie constraint functions
    # (hx)^2 + fx +c
    #this constraint is stated as (bp*x + cip)^2 - ep*x - d - y <= 0
    # the cost of y will drive it to be equal to (bp*x + cip)^2 - ep*x -d
    def turbine_y_consume(index):
        i, t = index
        return cvxpy.power(turbine_para[i].fundata["bp"]*turbine_xp_k[i,t],2)\
        + turbine_para[i].fundata["fp"]*turbine_xp_k[i,t]\
        + turbine_para[i].fundata["cp"]*turbine_s_k[i,t]\
        + cvxpy.power(turbine_para[i].fundata["bq"]*turbine_xq_k[i,t],2)\
        + turbine_para[i].fundata["fq"]*turbine_xq_k[i,t]\
        + turbine_para[i].fundata["cq"]*turbine_s_k[i,t]\
        - turbine_y[i,t] <= 0

    def turbine_xp_generate(index):
        i, t = index
        return turbine_xp[i,t] == cvxpy.sum(turbine_xp_k[i,t])

    def turbine_xq_generate(index):
        i, t = index
        return turbine_xq[i,t] == cvxpy.sum(turbine_xq_k[i,t])

    def turbine_xp_k_lower(index):
        i, t = index
        #individual lower bounds are non-zero
        return (turbine_para[i].lb)*turbine_s_k[i,t] <= turbine_xp_k[i,t]

    def turbine_xq_k_lower(index):
        i, t = index
        return (turbine_para[i].lb)*0.01*turbine_s_k[i,t] <= turbine_xq_k[i,t]

    def turbine_xp_k_upper(index):
        i, t = index
        return (turbine_para[i].ub)*turbine_s_k[i,t] >= turbine_xp_k[i,t] 
        #+ cvxpy.power(turbine_xq_k[i,t],2)

    def turbine_x_status(index):
        i, t = index
        #return turbine_s[i,t] == cvxpy.sum(turbine_s_k[i,t])
        return 1>=cvxpy.sum(turbine_s_k[i,t])

    # def turbine_start_status1(index):
    #     i, t = index[0], 0
    #     return turbine_start[i,t] >= turbine_s[i,1] - init_turbine[i].status

    def turbine_start_status(index):
        i, t = index
        return turbine_start[i,t] >= turbine_s[i,t] - turbine_s[i,t-1]

    def turbine_ramp1_up(index):
        i, t = index[0], 0
        return turbine_init[i] - turbine_para[i].ramp_rate <= turbine_xp[i,t]

    def turbine_ramp1_down(index):
        i, t = index[0], 0
        return turbine_xp[i,t] <= turbine_init[i] + turbine_para[i].ramp_rate

    def turbine_ramp_up(index):
        i, t = index
        return turbine_xp[i,t-1] + turbine_para[i].ramp_rate >= turbine_xp[i,t]

    def turbine_ramp_down(index):
        i, t = index
        return turbine_xp[i,t-1] - turbine_para[i].ramp_rate <= turbine_xp[i,t]

    def turbine_powerfactor_limit_upper(index):
        i, t = index
        return turbine_xq[i,t] <= turbine_xp[i,t]*0.2

    def turbine_powerfactor_limit_lower(index):
        i, t = index
        return turbine_xq[i,t] >= -turbine_xp[i,t]*0.2

    #turbines lock on time limit not defined in this configuration

    #diesel generator constraint functions

    # def dieselgen_y_consume(index):
    #     i, t = index
    #     return cvxpy.power(diesel_para[i].fundata["bp"]*dieselgen_xp_k[i,t],2)\
    #     + diesel_para[i].fundata["fp"]*dieselgen_xp_k[i,t]\
    #     + diesel_para[i].fundata["cp"]*dieselgen_s_k[i,t]\
    #     + cvxpy.power(diesel_para[i].fundata["bq"]*dieselgen_xq_k[i,t],2)\
    #     + diesel_para[i].fundata["fq"]*dieselgen_xq_k[i,t]\
    #     - diesel_para[i].fundata["cq"]*dieselgen_s_k[i,t]\
    #     - dieselgen_y[i,t] <= 0

    # def dieselgen_xp_generate(index):
    #     i, t = index
    #     return dieselgen_xp[i,t] == cvxpy.sum(dieselgen_xp_k[i,t])

    # def dieselgen_xq_generate(index):
    #     i, t = index
    #     return dieselgen_xq[i,t] == cvxpy.sum(dieselgen_xq_k[i,t])

    # def dieselgen_xp_k_lower(index):
    #     i, t = index
    #     #individual lower bounds are zero
    #     return diesel_para[i].lb*dieselgen_s_k[i,t] <= dieselgen_xp_k[i,t]

    # def dieselgen_xp_k_upper(index):
    #     i, t = index
    #     return diesel_para[i].ub*dieselgen_s_k[i,t] >= dieselgen_xp_k[i,t]

    # def dieselgen_x_status(index):
    #     i, t = index
    #     #return dieselgen_s[i,t]== cvxpy.sum(dieselgen_s_k[i,t])
    #     return 1>=cvxpy.sum(dieselgen_s_k[i,t])

    # def dieselgen_start_status(index):
    #     i, t = index
    #     return dieselgen_start[i,t] >= dieselgen_s[i,t] - dieselgen_s[i,t-1]

    # def dieselgen_ramp1_up(index):
    #     i, t = index[0], 0
    #     return dieselgen_xp[i,t] <= dieselgen_init[i] + diesel_para[i].ramp_rate

    # def dieselgen_ramp1_down(index):
    #     i, t = index[0], 0
    #     return dieselgen_init[i] - diesel_para[i].ramp_rate <= dieselgen_xp[i,t]

    # def dieselgen_ramp_up(index):
    #     i, t = index
    #     return dieselgen_xp[i,t-1] + diesel_para[i].ramp_rate >= dieselgen_xp[i,t]

    # def dieselgen_ramp_down(index):
    #     i, t = index
    #     return dieselgen_xp[i,t-1] - diesel_para[i].ramp_rate <= dieselgen_xp[i,t]

    # def dieselgen_powerfactor_limit_upper(index):
    #     i, t = index
    #     return dieselgen_xq[i,t] <= dieselgen_xp[i,t] * 0.2

    # def dieselgen_powerfactor_limit_lower(index):
    #     i, t = index
    #     return dieselgen_xq[i,t] >= -dieselgen_xp[i,t] * 0.2

    # boiler constraint functions

    def boiler_y_consume(index):
        i, t = index 
        return cvxpy.power(boiler_para[i].fundata["b"]*boiler_x_k[i,t], 2)\
        + boiler_para[i].fundata["f"]*boiler_x_k[i,t]\
        + boiler_para[i].fundata["c"]*boiler_s_k[i,t]\
        - boiler_y[i,t] <=0

    def boiler_x_generate(index):
        i, t = index
        return boiler_x[i,t] == cvxpy.sum(boiler_x_k[i,t])

    def boiler_x_k_lower(index):
        i, t = index
        return boiler_s_k[i,t]*boiler_para[i].lb <= boiler_x_k[i,t]

    def boiler_x_k_upper(index):
        i, t = index
        return boiler_para[i].ub*boiler_s_k[i,t] >= boiler_x_k[i,t]

    def boiler_x_status(index):
        i, t = index
        #return boiler_s[i,t] == cvxpy.sum(boiler_s_k[i,t])
        return 1>=cvxpy.sum(boiler_s_k[i,t])

    def boiler_start_status1(index):
        i, t = index[0], 0
        return boiler_start[i,t] >= boiler_s[i,1] - boiler_init[i].status

    def boiler_start_status(index):
        i, t = index
        return boiler_start[i,t] >= boiler_s[i,t] - boiler_s[i,t-1]

    def boiler_ramp1_up(index):
        i, t = index[0], 0
        return boiler_x[i,t] <= boiler_init[i] + boiler_para[i].ramp_rate
    def boiler_ramp1_down(index):
        i, t = index[0], 0
        return boiler_init[i] - boiler_para[i].ramp_rate <= boiler_x[i,t]

    def boiler_ramp_up(index):
        i, t = index
        return boiler_x[i,t-1] + boiler_para[i].ramp_rate >= boiler_x[i,t]

    def boiler_ramp_down(index):
        i, t = index
        return boiler_x[i,t-1] - boiler_para[i].ramp_rate <= boiler_x[i,t]

    # chiller constraints

    def chiller_yp_consume(index):
        i, t = index
        return cvxpy.power(chiller_para[i].fundata["bp"]*chiller_x_k[i,t], 2)\
        + chiller_para[i].fundata["fp"]*chiller_x_k[i,t]\
        + chiller_para[i].fundata["cp"]*chiller_s_k[i,t]\
        -chiller_yp[i,t] <= 0

    def chiller_yq_consume(index):
        i, t = index
        return chiller_yq[i,t] == chiller_yp[i,t]*0.05
        # return cvxpy.power(chiller_para[i].fundata["bq"]*chiller_x_k[i,t], 2)\
        # + chiller_para[i].fundata["fq"]*chiller_x_k[i,t]\
        # + chiller_para[i].fundata["cq"]*chiller_s_k[i,t] - chiller_yq[i,t] <= 0

    def chiller_x_generate(index):
        i, t = index
        return chiller_x[i,t] == cvxpy.sum(chiller_x_k[i,t])

    # def chiller_x_lower(index):
    #     i, t = index
    #     return chiller_para[i].lb *chiller_s[i,t] <= chiller_x[i,t]

    def chiller_x_k_lower(index):
        i, t = index
        return chiller_para[i].lb *chiller_s_k[i,t] <= chiller_x_k[i,t]
    
    def chiller_y_bound(index):
        i, t = index
        return chiller_para[i].ub * 5 * chiller_s_k[i,t] >= chiller_yp[i,t]

    def chiller_x_k_upper(index):
        i, t = index
        return chiller_para[i].ub * chiller_s_k[i,t] >= chiller_x_k[i,t]

    def chiller_x_status(index):
        i, t = index
        #return chiller_s[i,t] == cvxpy.sum(chiller_s_k[i,t])
        return 1>=cvxpy.sum(chiller_s_k[i,t])

    def chiller_start_status1(index):
        i, t = index[0], 0
        return chiller_start[i,t] >= chiller_s[i,1] - chiller_init[i].status

    def chiller_start_status(index):
        i, t = index
        return chiller_start[i,t] >= chiller_s[i,t] - chiller_s[i, t-1]

    def chiller_ramp1_up(index):
        i, t = index[0], 0
        return chiller_x[i,t] <= chiller_init[i] + chiller_para[i].ramp_rate

    def chiller_ramp1_down(index):
        i, t = index[0], 0
        return chiller_init[i] - chiller_para[i].ramp_rate <= chiller_x[i,t]

    def chiller_ramp_up(index):
        i, t = index
        return chiller_x[i, t-1] + chiller_para[i].ramp_rate >= chiller_x[i,t]

    def chiller_ramp_down(index):
        i, t = index
        return chiller_x[i, t-1] - chiller_para[i].ramp_rate <= chiller_x[i,t]


    # power dumping constraint
    def wasted_heat(index):
        t = index[0]
        return q_hru_heating_in[0,t] == turbine_y[0,t] - turbine_xp[0,t]/293.1 + turbine_y[1,t] - turbine_xp[1,t]/293.1 #-abs_y[0,t]

    def hru_limit(index):
        t = index[0]
        return q_hru_heating_out[0,t] <= a_hru * q_hru_heating_in[0,t]

    def no_slack_c(index):
        i, t = index
        return cool_unserve[i,t] == 0

    def no_slack_h(index):
        i,t = index
        return heat_unserve[i,t] == 0

    def no_slack_e(index):
        i, t = index
        return elec_unserve[i,t] == 0

    ###### DEFINE COMPONENT TYPES
    # electric grid
    tic = time.time()
    index_hour = (range(T),)
    index_nodes = range(n_nodes), range(T)
    ep_elecfromgrid = VariableGroup("ep_elecfromgrid", indexes=index_nodes, lower_bound_func=constant_zero) #real power from grid
    ep_electogrid = VariableGroup("ep_electogrid", indexes=index_nodes, lower_bound_func=constant_zero) #real power to the grid
    eq_elecfromgrid = VariableGroup("eq_elecfromgrid", indexes=index_nodes, lower_bound_func=constant_zero) #reactive power from grid
    eq_electogrid = VariableGroup("eq_electogrid", indexes=index_nodes, lower_bound_func=constant_zero) #reactive power from grid
    ep_gridstate = VariableGroup("ep_gridstate", indexes=index_nodes, is_binary_var=True)
    #eq_gridstate = VariableGroup("eq_gridstate", indexes=index_nodes, is_binary_var=True) # if you are importing or exporting at a node


    #dumping allowance
    elec_unserve = VariableGroup("elec_unserve", indexes=index_nodes, lower_bound_func=constant_zero)
    if n_boilers>0:
        heat_unserve = VariableGroup("heat_unserve", indexes=index_nodes, lower_bound_func=constant_zero)
        heat_dump = VariableGroup("heat_dump", indexes=index_nodes, lower_bound_func=constant_zero)
    if n_chillers>0:
        cool_unserve = VariableGroup("cool_unserve", indexes=index_nodes, lower_bound_func=constant_zero)
        #cool_dump = VariableGroup("cool_dump", indexes=index_nodes, lower_bound_func=constant_zero)
    if allow_thermal_slack==False:
        add_constraint("no_slack_h", index_nodes, no_slack_h)
        add_constraint("no_slack_c", index_nodes, no_slack_c)
        add_constraint("no_slack_e", index_nodes, no_slack_e)


    #turbines: # fuel cells are considered turbines
    index_turbines = range(n_turbines), range(T)
    turbine_y = VariableGroup("turbine_y", indexes =index_turbines, lower_bound_func=constant_zero) #  fuel use
    turbine_xp = VariableGroup("turbine_xp", indexes=index_turbines, lower_bound_func=constant_zero)  #  real power output
    turbine_xq = VariableGroup("turbine_xq", indexes=index_turbines, lower_bound_func=constant_zero)  #  reactive power output
    turbine_xp_k = VariableGroup("turbine_xp_k", indexes=index_turbines, pieces=turbine_pieces, lower_bound_func=constant_zero) #  power outputs from all piecewise parts
    turbine_xq_k = VariableGroup("turbine_xq_k", indexes=index_turbines,pieces=turbine_pieces) #  power outputs from all piecewise parts
    turbine_s_k = VariableGroup("turbine_s_k", indexes=index_turbines, is_binary_var=True, pieces=turbine_pieces) #  states from all pieceswise parts
    #turbine_s = VariableGroup("turbine_s", indexes=index_turbines, is_binary_var=True)# unit commitment of turbine

    #diesel generators:
    # index_dieselgen = range(n_dieselgen), range(T)
    # dieselgen_y = VariableGroup("dieselgen_y", indexes=index_dieselgen, lower_bound_func=constant_zero) #fuel use
    # dieselgen_xp = VariableGroup("dieselgen_xp", indexes=index_dieselgen, lower_bound_func=constant_zero) # real power output
    # dieselgen_xq = VariableGroup("dieselgen_xq", indexes=index_dieselgen, lower_bound_func=constant_zero) # reactive power output
    # dieselgen_xp_k = VariableGroup("dieselgen_xp_k", indexes=index_dieselgen, pieces=diesel_pieces, lower_bound_func=constant_zero) # power outputs from all piecewise parts
    # dieselgen_xq_k = VariableGroup("dieselgen_xq_k", indexes=index_dieselgen, pieces=diesel_pieces) # power outputs from all piecewise parts
    # dieselgen_s_k = VariableGroup("dieselgen_s_k", indexes=index_dieselgen, is_binary_var=True, pieces=diesel_pieces) # states from all piecewise pats
    #dieselgen_s = VariableGroup("dieselgen_s", indexes=index_dieselgen, is_binary_var = True) #unit commitment


    #boilers:
    index_boilers = range(n_boilers), range(T)
    boiler_y = VariableGroup("boiler_y", indexes=index_boilers, lower_bound_func=constant_zero) #  fuel use from boiler
    boiler_x = VariableGroup("boiler_x", indexes=index_boilers, lower_bound_func=constant_zero) #  heat output from boiler
    boiler_x_k = VariableGroup("boiler_x_k", indexes=index_boilers, pieces=boiler_pieces, lower_bound_func = constant_zero) #  heat output from each portion of the piecewise fit
    boiler_s_k = VariableGroup("boiler_s_k", indexes=index_boilers, is_binary_var=True, pieces=boiler_pieces) #  unit commitment for each portion of the piecewise efficiency fit
    #boiler_s = VariableGroup("boiler_s", indexes = index_boilers, is_binary_var = True) #unit commitment

    #chillers
    index_chiller = range(n_chillers), range(T)
    chiller_x = VariableGroup("chiller_x", indexes = index_chiller, lower_bound_func = constant_zero) #  cooling power output
    chiller_yp = VariableGroup("chiller_yp", indexes = index_chiller, lower_bound_func = constant_zero) #  real electric power demand
    chiller_yq = VariableGroup("chiller_yq", indexes = index_chiller, lower_bound_func = constant_zero) #  reactive electric power demand
    chiller_x_k = VariableGroup("chiller_x_k", indexes = index_chiller, pieces=chiller_pieces, lower_bound_func=constant_zero) #  cooling output from all piecewise parts
    chiller_s_k = VariableGroup("chiller_s_k", indexes=index_chiller, is_binary_var = True, pieces=chiller_pieces) #  unit commitment for piecewise sections
    #chiller_s = VariableGroup("chiller_s", indexes=index_chiller, is_binary_var=True) #unit commitment


    #storage
    #electric storage
    if n_e_storage>0:
        index_e_storage = range(n_e_storage), range(T)
        e_storage_disch = VariableGroup("e_storage_disch", indexes=index_e_storage, lower_bound_func = constant_zero)
        e_storage_ch = VariableGroup("e_storage_ch", indexes = index_e_storage, lower_bound_func = constant_zero)
        e_storage_state = VariableGroup("e_storage_state", indexes = index_e_storage, lower_bound_func =constant_zero)
    #cold water tank or other cold energy storage
    index_c_storage = range(n_c_storage), range(T)
    c_storage_disch = VariableGroup("c_storage_disch", indexes = index_c_storage, lower_bound_func = constant_zero)
    c_storage_ch = VariableGroup("c_storage_ch", indexes = index_c_storage, lower_bound_func = constant_zero)
    c_storage_state = VariableGroup("c_storage_state", indexes = index_c_storage, lower_bound_func = constant_zero)
    #hot water tank or other hot energy storage
    if n_h_storage>0:
        index_h_storage = range(n_h_storage), range(T)
        h_storage_disch = VariableGroup("h_storage_disch", indexes=index_h_storage, lower_bound_func=constant_zero)
        h_storage_ch = VariableGroup("h_storage_ch", indexes = index_h_storage, lower_bound_func = constant_zero)
        h_storage_state = VariableGroup("h_storage_state", indexes = index_h_storage, lower_bound_func = constant_zero)

    #nodal network
    #voltage is split into x, y, z
    #x_m = v_m^2 and is therefore positive
    #y_mn = v_m*v_n*cos(theta_mn)
    #z_mn = v_m*v_n*sin(theta_mn)
    index_e_nodes = range(n_e_nodes), range(T)
    index_e_lines = range(n_e_lines), range(T)
    x_m = VariableGroup("x_m", indexes = index_e_nodes, lower_bound_func = constant_zero)
    y_mn = VariableGroup("y_mn", indexes = index_e_lines, lower_bound_func = constant_zero) 
    z_mn = VariableGroup("z_mn", indexes = index_e_lines)
    #heat network
    index_h_nodes = range(n_h_nodes), range(T)
    index_h_lines = range(n_h_lines), range(T)
    h_mn = VariableGroup("h_mn", indexes = index_h_lines, lower_bound_func = constant_zero)
    #cooling network
    index_c_lines = range(n_c_lines), range(T)
    c_mn = VariableGroup("c_mn", indexes = index_c_lines, lower_bound_func = constant_zero)

    #define utility costs
    pelec_cost = [find_utility_pricing(date_stamp) for date_stamp in date_range]
    qelec_cost = np.multiply(pelec_cost,5)
    pselback_rate = np.multiply(pelec_cost,0)
    qselback_rate = np.multiply(qelec_cost, 0)
    gas_rate = [find_gas_pricing(date_stamp) for date_stamp in date_range]
    #diesel_rate = [find_diesel_pricing(date_stamp) for date_stamp in date_range]

    #forecast generation and demand
    forecast = TestData()
    forecast.demand.h = np.zeros((len(network),T))
    forecast.demand.c = np.zeros((len(network),T))
    setattr(forecast.demand, 'ep', np.zeros((len(network), T)))
    setattr(forecast.demand, 'eq', np.zeros((len(network), T)))
    setattr(forecast, 'renew', np.zeros((len(network),T)))
    i = 0
    for node in network:
        #if not node.electrical.load == []:
        ep_demand = [find_demand(date_stamp, 'e', n=node.electrical.load) for date_stamp in date_range]
        forecast.demand.ep[i,:] = ep_demand #np.multiply(ep_demand, 1/n_nodes)
        forecast.demand.eq[i,:] = np.multiply(ep_demand, 0.05)#assume high power factor for now
        if not node.district_heat.load == None:
            h = [find_demand(date_stamp, 'h', n =node.district_heat.load) for date_stamp in date_range]
            #h = [load[0] for load in h]
            forecast.demand.h[i,:] = np.multiply(h, 1)
        if not node.district_cooling.load == None:
            c = [find_demand(date_stamp, 'c', n=node.district_cooling.load) for date_stamp in date_range]
            #c = [load[0] for load in c]
            forecast.demand.c[i,:] = np.multiply(c,1)
        forecast.renew[i,:] = np.array([find_solar_forecast(date_stamp, n=i) for date_stamp in date_range])
        i +=1


    toc = time.time()-tic
    print('Variables '+str(toc))

    ######## OBJECTIVE FUNCTION

    tic = time.time()
    objective_components = []

    for var, _lambda in zip(ep_elecfromgrid[RANGE]*p_base, pelec_cost):
        objective_components.append(var * _lambda)

    for var, _lambda in zip(eq_elecfromgrid[RANGE]*p_base, qelec_cost):
        objective_components.append(var * _lambda)

    # for var, _lambda in zip(ep_electogrid[(RANGE,)]*p_base, pselback_rate):
    #     objective_components.append(var * _lambda)

    # for var, _lambda in zip(eq_electogrid[(RANGE,)]*p_base, qselback_rate):
    #     objective_components.append(var * _lambda)

    for i in range(n_turbines):
        for var, state, _lambda in zip(turbine_y[i, RANGE]*p_base, turbine_s_k[i,RANGE], gas_rate):
            objective_components.append(var * _lambda)
        # for var in turbine_start[i, RANGE]:
        #     objective_components.append(var * turbine_para[i].start_cost)

    # for i in range(n_dieselgen):
    #     for var, _lambda in zip(dieselgen_y[(i, RANGE)]*p_base, diesel_rate):
    #         objective_components.append(var * _lambda)
    #     # for var in dieselgen_start[i, RANGE]:
    #     #     objective_components.append(var * diesel_para[i].start_cost)

    for i in range(n_boilers):
        for var, _lambda in zip(boiler_y[i, RANGE]*h_base, gas_rate):
            objective_components.append(var * _lambda)
        # for var in boiler_start[i, RANGE]:
        #     objective_components.append(var * boiler_para[i].start_cost)

    stored_water_rate = np.mean(gas_rate)/(0.5*6.5) # the gas rate times approx gen efficiency times chiller COP
    for i in range(n_c_storage):
        objective_components.append((c_storage_state[i,0]-c_storage_state[i,T-1])*stored_water_rate)

    #only penalize unserved demand
    if allow_thermal_slack:
        for group in (heat_unserve, cool_unserve, elec_unserve):
            for var in group[RANGE]:
                objective_components.append(var * bigM)

    #toc = time.time()-tic
    #print('objective function '+str(toc))

    ######## ADD CONSTRAINTS
    #tic = time.time()
    index_without_first_hour = (range(1,T),)

    # add turbine constraints 
    index_turbine = (range(n_turbines),)
    add_constraint("turbine_y_consume", index_turbine + index_hour, turbine_y_consume) #False
    add_constraint("turbine_xp_generate", index_turbine + index_hour, turbine_xp_generate) #True
    add_constraint("turbine_xq_generate", index_turbine + index_hour, turbine_xq_generate)
    add_constraint("turbine_xp_k_lower", index_turbine + index_hour, turbine_xp_k_lower)
    add_constraint("turbine_xp_k_upper", index_turbine + index_hour, turbine_xp_k_upper)
    add_constraint("turbine_x_status", index_turbine + index_hour, turbine_x_status)
    #add_constraint("turbine_start_status1", index_turbine, turbine_start_status1)
    #add_constraint("turbine_start_status", index_turbine + index_without_first_hour, turbine_start_status)
    add_constraint("turbine_ramp1_up", index_turbine, turbine_ramp1_up)
    add_constraint("turbine_ramp1_down", index_turbine, turbine_ramp1_down)
    add_constraint("turbine_ramp_up", index_turbine + index_without_first_hour, turbine_ramp_up)
    add_constraint("turbine_ramp_down", index_turbine + index_without_first_hour, turbine_ramp_down)
    add_constraint("turbine_powerfactor_limit_upper", index_turbine+index_hour, turbine_powerfactor_limit_upper)
    add_constraint("turbine_powerfactor_limit_lower", index_turbine+index_hour, turbine_powerfactor_limit_lower)
    #add_constraint("turbines_lock_on1", index_turbine, turbines_lock_on1)

    # # add diesel constraints
    # index_diesel = (range(n_dieselgen),)
    # add_constraint("dieselgen_y_consume", index_diesel + index_hour, dieselgen_y_consume)
    # add_constraint("dieselgen_xp_generate", index_diesel + index_hour, dieselgen_xp_generate)
    # add_constraint("dieselgen_xq_generate", index_diesel + index_hour, dieselgen_xq_generate)
    # add_constraint("dieselgen_xp_k_lower", index_diesel + index_hour, dieselgen_xp_k_lower)
    # add_constraint("dieselgen_xp_k_upper", index_diesel + index_hour, dieselgen_xp_k_upper)
    # add_constraint("dieselgen_x_status", index_diesel + index_hour, dieselgen_x_status)
    # #add_constraint("dieselgen_start_status", index_diesel + index_without_first_hour, dieselgen_start_status)
    # add_constraint("dieselgen_ramp1_up", index_diesel, dieselgen_ramp1_up)
    # add_constraint("dieselgen_ramp1_down", index_diesel, dieselgen_ramp1_down)
    # add_constraint("dieselgen_ramp_up", index_diesel + index_without_first_hour, dieselgen_ramp_up)
    # add_constraint("dieselgen_ramp_down", index_diesel + index_without_first_hour, dieselgen_ramp_down)
    # add_constraint("dieselgen_powerfactor_limit_upper", index_diesel + index_hour, dieselgen_powerfactor_limit_upper)
    # add_constraint("dieselgen_powerfactor_limit_lower", index_diesel + index_hour, dieselgen_powerfactor_limit_lower)

    # add boiler constraints
    index_boiler = (range(n_boilers),)
    add_constraint("boiler_y_consume", index_boiler + index_hour, boiler_y_consume)
    add_constraint("boiler_x_generate", index_boiler + index_hour, boiler_x_generate)
    add_constraint("boiler_x_k_lower", index_boiler + index_hour, boiler_x_k_lower)
    add_constraint("boiler_x_k_upper", index_boiler + index_hour, boiler_x_k_upper)
    add_constraint("boiler_x_status", index_boiler + index_hour, boiler_x_status)
    #add_constraint("boiler_start_status1", index_boiler + index_hour, boiler_start_status1)
    #add_constraint("boiler_start_status", index_boiler + index_without_first_hour, boiler_start_status)
    add_constraint("boiler_ramp1_up", index_boiler, boiler_ramp1_up)
    add_constraint("boiler_ramp1_down", index_boiler, boiler_ramp1_down)
    add_constraint("boiler_ramp_up", index_boiler + index_without_first_hour, boiler_ramp_up)
    add_constraint("boiler_ramp_down", index_boiler + index_without_first_hour, boiler_ramp_down)

    #add chiller constriants
    index_chiller = (range(n_chillers),)
    add_constraint("chiller_yp_consume", index_chiller + index_hour, chiller_yp_consume)
    add_constraint("chiller_yq_consume", index_chiller + index_hour, chiller_yq_consume)
    add_constraint("chiller_x_generate", index_chiller + index_hour, chiller_x_generate)
    add_constraint("chiller_x_k_lower", index_chiller + index_hour, chiller_x_k_lower)
    add_constraint("chiller_x_k_upper", index_chiller + index_hour, chiller_x_k_upper)
    add_constraint("chiller_y_bound", index_chiller + index_hour, chiller_y_bound)
    add_constraint("chiller_x_status", index_chiller + index_hour, chiller_x_status)
    #add_constraint("chiller_start_status1", index_chiller, chiller_start_status1)
    #add_constraint("chiller_start_status", index_chiller + index_without_first_hour, chiller_start_status)
    add_constraint("chiller_ramp1_up", index_chiller, chiller_ramp1_up)
    add_constraint("chiller_ramp1_down", index_chiller, chiller_ramp1_down)
    add_constraint("chiller_ramp_up", index_chiller + index_without_first_hour, chiller_ramp_up)
    add_constraint("chiller_ramp_down", index_chiller + index_without_first_hour, chiller_ramp_down)
    #add_constraint("chiller_power_ratio", index_chiller+ index_hour, chiller_power_ratio)

    #add absorption chillers
    # index_abs = (range(n_abs,),)
    # add_constraint("abs_y_consume", index_abs + index_hour, abs_y_consume)
    # add_constraint("abs_x_generate", index_abs + index_hour, abs_x_generate)
    # add_constraint("abs_x_lower", index_abs + index_hour, abs_x_lower)
    # add_constraint("abs_x_k_lower", index_abs + index_hour, abs_x_k_lower)
    # add_constraint("abs_x_k_upper", index_abs + index_hour, abs_x_k_upper)
    # add_constraint("abs_x_status", index_abs + index_hour, abs_x_status)
    # #add_constraint("abs_start_status1", index_abs, abs_start_status1)
    # #add_constraint("abs_start_status", index_abs + index_without_first_hour, abs_start_status)
    # add_constraint("abs_ramp1_up", index_abs, abs_ramp1_up)
    # add_constraint("abs_ramp1_down", index_abs, abs_ramp1_down)
    # add_constraint("abs_ramp_up", index_abs + index_without_first_hour, abs_ramp_up)
    # add_constraint("abs_ramp_down", index_abs + index_without_first_hour, abs_ramp_down)

    #wasted heat
    #add_constraint("wasted_heat", index_hour, wasted_heat)
    #add_constraint("HRUlimit", index_hour, HRUlimit)


    # add storage constraints
    if n_e_storage>0:
        index_e_storage = (range(n_e_storage),)
        add_constraint("e_storage_init", index_e_storage, e_storage_init)
        add_constraint("e_storage_state_constraint", index_e_storage + index_without_first_hour, e_storage_state_constraint)
        #add_constraint("e_storage_end", index_e_storage, e_storage_end)

    if n_h_storage>0:
        index_h_storage = (range(n_h_storage),)
        add_constraint("h_storage_init", index_h_storage, h_storage_init)
        add_constraint("h_storage_state_constraint", index_h_storage + index_without_first_hour, h_storage_state_constraint)
        #add_constraint("h_storage_end", index_h_storage, h_storage_end)

    if n_c_storage>0:
        index_c_storage = (range(n_c_storage),)
        add_constraint("c_storage_init", index_c_storage, c_storage_init)
        add_constraint("c_storage_state_constraint", index_c_storage + index_without_first_hour, c_storage_state_constraint)
        #add_constraint("c_storage_end", index_c_storage, c_storage_end)


    # add equality constraints for supply and demand
    #for m in range(n_nodes):
    m_index = (e_nodes,)
    h_index = (h_nodes,)
    c_index = (c_nodes,)
    add_constraint("electric_p_balance", index_hour + m_index, electric_p_balance)
    add_constraint("electric_q_balance", index_hour + m_index, electric_q_balance)
    add_constraint("cool_balance", index_hour + c_index, cool_balance)
    add_constraint("heat_balance", index_hour + h_index, heat_balance)

    # add line and voltage limits
    #for m in range(n_nodes):
    add_constraint("power_import_state", index_hour + m_index, power_import_state)
    add_constraint("power_export_state", index_hour + m_index, power_export_state)
    add_constraint("voltage_limit_upper", index_hour + m_index, voltage_limit_upper)
    add_constraint("voltage_limit_lower", index_hour + m_index, voltage_limit_lower)
    #add_constraint("y_voltage_limit_upper", index_hour+ n_index, y_voltage_limit_upper)
    #add_constraint("y_voltage_limit_lower", index_hour+n_index, y_voltage_limit_lower)
    #add_constraint("z_voltage_limit_upper", index_hour+n_index, z_voltage_limit_upper)
    #add_constraint("z_voltage_limit_lower", index_hour+n_index, z_voltage_limit_lower)#this one makes it switch from primal unbounded, to primal infeasible
    m_connected_index = (e_connected_nodes,)
    add_constraint("current_limit", index_hour + m_connected_index, current_limit)
    add_constraint("y_mn_equality", index_hour + m_connected_index, y_mn_equality)
    add_constraint("z_mn_inequality", index_hour + m_connected_index, z_mn_inequality)

    # add variable subsitution constraint
    #for m in range(n_nodes):
    add_constraint("electric_interrelation", index_hour + m_connected_index, electric_interrelation) # not sure this constraint is needed
    add_constraint("line_heat", index_hour + h_index, line_heat)
    add_constraint("line_cooling", index_hour + c_index, line_cooling)

    #toc = time.time()-tic
    #print('add constraints: '+ str(toc))

    print('problem parameters loaded')


    ######## SOLVE FINAL PROBLEM
    objective = cvxpy.Minimize(cvxpy.sum(objective_components))
    constraints_list = [x[0] for x in constraints]
    prob = cvxpy.Problem(objective, constraints_list)
    print('problem created, solving problem')

    tic = time.time()
    # if v_iters>1:
    #     result = prob.solve(solver='GUROBI',verbose=True, NumericFocus=1, IterationLimit=500)#  
    result = prob.solve(solver='GUROBI', NumericFocus=3)#, verbose=True)#NumericFocus=1, IterationLimit=10,  #solver = 'ECOS_BB')#verbose = True, , NumericFocus=3 warmstart = True, 


    toc = time.time()-tic
    print('optimal cost: '+ str(result))
    print('time step: ' + str(timestep))
    #if result != float('inf'):
    #    sort_solution(prob, var_name_list,T, forecast, timestep)#print(prob._solution)
    print('problem solved in '+str(toc)+'seconds')

    # find error in voltage convergence
    #voltage_error = [x_m[i,t].value[0] - x_n[i,t] for i,t in product(range(n_nodes), range(T))]
    voltage_error = np.zeros((n_e_nodes, T))
    pid_error = np.zeros((n_e_nodes, T))
    # if v_iters == 1:
    #     k = 0
    # else:
    #     k = 0.3

    for m in range(n_e_nodes):
        for t in range(T):
            # calculate error
            voltage_error[m,t] = x_m[m,t].value[0] - x_n[m,t]
            # save for next iteration
            # it is used as an upper limit, so try to prevent infeasibilities
            #x_n[m,t] = np.sqrt(x_m[m,t].value[0]*x_n[m,t]) +0.001
            #pid_error[m,t] = voltage_error[m,t]#(np.sqrt(x_m[m,t].value[0]*x_n[m,t])-x_n[m,t])#

            # if the sign of the error changes, then you have overstepped, reduce step size
            # if (voltage_error[m,t] < 0 and pid_error_last[m,t] > 0) or  (voltage_error[m,t] > 0 and pid_error_last[m,t] < 0):
            #     pid_error[m,t] = -2/3*copy(pid_error_last[m,t])
            #     #voltage_error[m,t] = copy(pid_error_last[m,t])#0
            # else:
            # if voltage_error[m,t]<0:
            #     k = -.8
            # else:
            #     k = .8
            x_n[m,t] = x_n[m,t] + voltage_error[m,t]#*(1-0.01/3)
            # pid_error[m,t] = k*((voltage_error[m,t])**2)#+k/2*np.abs(voltage_error[m,t])#pid_error_last[m,t]*.25

            # #x_n = x_n_previous + difference*gain
            # xni = x_n[m,t] + pid_error[m,t]#voltage_error[m,t]*k #+ (pid_error_last[m,t]-pid_error[m,t])*0.1
            # xni = min(xni, (1.0 + voltage_deviation)**2)
            # xni = max(xni, (1.0 - voltage_deviation)**2)
            # x_n[m,t] = xni
            #x_n[m,t] = x_n[m,t] + pid_error[m,t]*0.75 + (pid_error_last[m,t]-pid_error[m,t])*0.2
            #0.85 becomes unstable at iteration 138
            #x_n[m,t] = np.sqrt(x_m[m,t].value[0]*x_n[m,t])
            #x_n[m,t] = x_n[m,t] + (np.sqrt(x_m[m,t].value[0]*x_n[m,t])-x_n[m,t])*0.5
            

    #if values have converged, save values
    if np.all(abs(voltage_error)<= 0.005) or v_iters>=19:
        pid_error = np.zeros((n_e_nodes, T))
        if v_iters >=19:
            print('no convergence after 20 iterations, start clean for next timestep')
            x_n = np.ones((n_e_nodes,T))
        else:
            print('voltage value convergence after '+ str(v_iters) + ' iterations')
        v_iters = 0
        print('parsing and saving solution')
        tic = time.time()
        ######## SORT PROBLEM SOLUTION
        #create a csv of the results
        # make the first row, the name list
        filename = 'command_output_wsu.csv'
        field_names = []
        if timestep==0:
            open_method = 'w'
            # for i in range(len(var_name_list)):
            #     var_name = var_name_list[i]
            #     split_name = var_name.split('_')
            #     var_name = var_name.split(split_name[-2])[0][:-1]
            #     j = int(split_name[-2])
            #     t = int(split_name[-1])

            # with open(filename, 'w', encoding='utf-8') as logfile:

        
        else:
            open_method = 'a'
        
        with open(filename, open_method) as logfile:
            values = {}
            for i in range(len(var_name_list)):
                var_name = var_name_list[i]
                split_name = var_name.split('_')
                var_name = var_name.split(split_name[-2])[0][:-1]
                j = int(split_name[-2])
                t = int(split_name[-1])
                field_name = var_name+'_'+str(j)
                # get numeric value
                var_val = eval(var_name)[j,t]
                if var_val.attributes['boolean']:
                    var_val = var_val.value
                elif var_val.value == None:
                    var_val = 0
                else:
                    var_val = var_val.value[0]
                # add to entry
                if field_name in values:
                    values[field_name].append(var_val)
                else:
                    field_names.append(field_name)
                    values[field_name] = [var_val]
            field_names.append('solar')
            field_names.append('ep_demand')
            field_names.append('eq_demand')
            field_names.append('h_demand')
            field_names.append('c_demand')
            values['solar'] = sum(forecast.renew)
            values['ep_demand'] = sum(forecast.demand.ep)
            values['eq_demand'] = sum(forecast.demand.eq)
            values['h_demand'] = sum(forecast.demand.h)
            values['c_demand'] = sum(forecast.demand.c)
            
            logger = csv.DictWriter(logfile, fieldnames = field_names, lineterminator = '\n')
            if timestep==0:
                logger.writeheader()
            # else:
            #     logger = csv.writer(logfile)
            for t in range(T):
                values_by_row = {}
                for key, value in values.items():
                    values_by_row[key] = value[t]
                #if timestep==0:
                logger.writerow(values_by_row)
                #else:
                    
            


        # for i in range(len(var_name_list)):
        #     var_name = var_name_list[i]
        #     # make the first row, the name list
        #     if timestep == 0:
        #         with open(output_filename, 'a') as outfile:
        #             outfile.write(var_name)
        #     split_name = var_name.split('_')
        #     var_name = var_name.split(split_name[-2])[0][:-1]
        #     j = int(split_name[-2])
        #     t = int(split_name[-1])
        #     # if it's a new timestep, make a new line
        #     if :
        #         # add the date 
        #         with open(output_filename, 'a') as outfile:
        #             outfile.wrtie(val)
        #         val = []

        #     var_val = eval(var_name[j,t])
        #     if var_val.attributes['boolean']:
        #         var_val = var_val.value
        #     elif var_val.value == None:
        #         var_val = 0
        #     else:
        #         var_val = var_val.value[0]
        #     val.append(var_val)
        # # also append the demand values
        # val.append(sum(forecast.renew[:,t]))
        # val.append(sum(forecast.demand.ep[:,t]))


        # create xlsx of results
        # if result != float('inf'):
        #     output_filename = 'command_output_wsu.xlsx'
        #     if timestep == 0:
        #         book = xlsxwriter.Workbook(output_filename)
        #         sheet1 = book.add_worksheet("Dispatch")
        #         n_col=-1
        #     else:
        #         book = openpyxl.load_workbook(output_filename)
        #         sheet1 = book.get_sheet_by_name("Dispatch")
        #         n_col=0
        #     #temp_vars = prob.solution.primal_vars
        #     j = 0
        #     t = 0
        #     start_row = timestep*T+1 #the row to start on
        #     for i in range(len(var_name_list)):
        #         var_name = var_name_list[i]
        #         split_name = var_name.split('_')
        #         var_name = var_name.split(split_name[-2])[0][:-1]
        #         j = int(split_name[-2])
        #         t = int(split_name[-1])
        #         #if you got to a new variable add a new heading, column
        #         if t == 0:
        #             n_col = n_col+1 
        #             if timestep == 0:
        #                 sheet1.write(t,n_col, var_name+'_'+str(j))
        #         n_row = t+start_row
        #         #write the value in the sheet
        #         var = eval(var_name)[j,t]
        #         if var.attributes['boolean']:
        #             val = var.value
        #         elif var.value == None:
        #             val = 0
        #         else:
        #             val = var.value[0]
        #         if timestep==0:
        #             sheet1.write(n_row,n_col, val)
        #         else:
        #             _ = sheet1.cell(column=n_col, row=n_row, value=val)
        #     # put renewable production and demand at the end of the sheet
        #     if t==0 and timestep==0:
        #         sheet1.write(0,n_col+1, 'renewable_gen')
        #         sheet1.write(0,n_col+2, 'demand_ep')
        #         sheet1.write(0,n_col+3, 'demand_eq')
        #         sheet1.write(0,n_col+4, 'demand_h')
        #         sheet1.write(0,n_col+5, 'demand_c')
        #         sheet1.write(0,n_col+6, 'datestamp')
        #         sheet1.write(n_row,n_col+6, date_range[0].strftime("%m%d%Y, %H:%M:%S"))
        #     elif t==0:
        #         _ = sheet1.cell(column=n_col+6, row=n_row, value = date_range[0].strftime("%m%d%Y, %H:%M:%S"))

        #     if timestep==0:
        #         for t in range(T):
        #             sheet1.write(t+start_row, n_col+1, sum(forecast.renew[:,t]))
        #             sheet1.write(t+start_row, n_col+2, sum(forecast.demand.ep[:,t]))
        #             sheet1.write(t+start_row, n_col+3, sum(forecast.demand.eq[:,t]))
        #             sheet1.write(t+start_row, n_col+4, sum(forecast.demand.h[:,t]))
        #             sheet1.write(t+start_row, n_col+5, sum(forecast.demand.c[:,t]))
        #         book.close()
        #     else:
        #         for t in range(T):
        #             _ = sheet1.cell(column=n_col+1, row=start_row+t, value=sum(forecast.renew[:,t]))
        #             _ = sheet1.cell(column=n_col+2, row=start_row+t, value=sum(forecast.demand.ep[:,t]))
        #             _ = sheet1.cell(column=n_col+3, row=start_row+t, value=sum(forecast.demand.eq[:,t]))
        #             _ = sheet1.cell(column=n_col+4, row=start_row+t, value=sum(forecast.demand.h[:,t]))
        #             _ = sheet1.cell(column=n_col+5, row=start_row+t, value=sum(forecast.demand.c[:,t]))
        #         _ = sheet1.cell(column=n_col+6, row=start_row, value = date_range[0].strftime("%m%d%Y, %H:%M:%S"))
        #         book.save(output_filename)
        toc = time.time()-tic
        print('time for parsing '+ str(toc))

    del prob
    return v_iters, x_n, pid_error


for t in range(timesteps):
    v_iters = 1
    pid_error = np.zeros((n_e_nodes, T))
    # allow up to ten iterations before just acceptig the last iteration as close enough
    while v_iters>0 and v_iters<20:
        constraints = []
        var_name_list = []
        # if you are on a subsequent timestep, try the voltages from the last timestep before starting over
        # at the maximum voltage deviation. This helps reduce iterations, but may cause an overshoot in later timesteps
        if v_iters == 1 and t>0:
            # try:
            v_iters, x_n, pid_error = run_horizon(t, v_iters, x_n, pid_error)
            # if the voltages from last timestep don't work, start from the maximum voltage deviation
            # except:
            #     x_n = np.ones((n_e_nodes,T))*(1+voltage_deviation)**2
            #     v_iters, x_n, pid_error = run_horizon(t, v_iters, x_n, pid_error)
        else: 
            v_iters, x_n, pid_error = run_horizon(t, v_iters, x_n, pid_error)
        # increment the iteration
        if v_iters>0:
            v_iters = v_iters+1
    
    if v_iters==20:
        error_string = 'Error: voltage iteration limit exceeded, continuing with best approximation for timestep '+str(t)
        print(error_string)

    # update x_n for next timestep
    for n in range(n_e_nodes):
        x_n_end = x_n[n][0]
        x_n[n][:-1] = x_n[n][1:]
        x_n[n][-1] = x_n_end
        #x_n[n][-1] = 1.0

    # update dates
    start_date = start_date + datetime.timedelta(hours=1)
    date_range = [start_date+datetime.timedelta(hours=i) for i in range(T)]
    #update initial conditions
    # read previous horizon's entry
    n_row = t*T+1
    # get the header indexing on the first iteration
    if t == 0:
        with open('command_output_wsu.csv', 'r') as filename:
            logger = csv.reader(filename)
            headers = next(logger, None)
        t_init_i = headers.index('turbine_xp_0')
        b_init_i = headers.index('boiler_x_0')
        c_init_i = headers.index('chiller_x_0')
        cws_init_i = headers.index('c_storage_state_0')

    with open('command_output_wsu.csv', 'r') as filename:
        logger = csv.reader(filename)
        # skip all previous rows until you get to the initial condition row
        for i in range(n_row):
            next(logger)
        row = next(logger)
        turbine_init[0] = float(row[t_init_i])
        turbine_init[1] = float(row[t_init_i+1])
        turbine_init[2] = float(row[t_init_i+2])
        turbine_init[3] = float(row[t_init_i+3])
        boiler_init[0] = float(row[b_init_i])
        boiler_init[1] = float(row[b_init_i+1])
        boiler_init[2] = float(row[b_init_i+2])
        boiler_init[3] = float(row[b_init_i+3])
        boiler_init[4] = float(row[b_init_i+4])
        chiller_init[0] = float(row[c_init_i])
        chiller_init[1] = float(row[c_init_i+1])
        chiller_init[2] = float(row[c_init_i+2])
        chiller_init[3] = float(row[c_init_i+3])
        chiller_init[4] = float(row[c_init_i+4])
        chiller_init[5] = float(row[c_init_i+5])
        chiller_init[6] = float(row[c_init_i+6])
        chiller_init[7] = float(row[c_init_i+7])
        chiller_init[8] = float(row[c_init_i+8])
        c_storage0[0] = float(row[cws_init_i])

    # sheet1 = pd.read_excel('command_output_wsu.xlsx', sheet_name='Dispatch')
    # turbine_init[0] = sheet1['turbine_xp_0'][n_row]
    # turbine_init[1] = sheet1['turbine_xp_1'][n_row]
    # #dieselgen_init[0] = sheet1['dieselgen_xp_0'][n_row]
    # boiler_init[0] = sheet1['boiler_x_0'][n_row]
    # boiler_init[1] = sheet1['boiler_x_1'][n_row]
    # boiler_init[2] = sheet1['boiler_x_2'][n_row]
    # boiler_init[3] = sheet1['boiler_x_3'][n_row]
    # chiller_init[0] = sheet1['chiller_x_0'][n_row]
    # chiller_init[1] = sheet1['chiller_x_1'][n_row]
    # chiller_init[2] = sheet1['chiller_x_2'][n_row]
    # chiller_init[3] = sheet1['chiller_x_3'][n_row]
    # chiller_init[4] = sheet1['chiller_x_4'][n_row]
    # chiller_init[5] = sheet1['chiller_x_5'][n_row]
    # chiller_init[6] = sheet1['chiller_x_6'][n_row]
    # chiller_init[7] = sheet1['chiller_x_7'][n_row]
    # chiller_init[8] = sheet1['chiller_x_8'][n_row]
    # c_storage0[0] = sheet1['c_storage_state_0'][n_row]
    




        